"""
validation/brats_benchmark.py — Benchmark BraTS.

═══════════════════════════════════════════════════════════════════════════
PRINCIPIO FONDAMENTALE: nessuna metrica deve guardare il ground truth per
decidere come mappare i cluster. Il mapping è fissato A PRIORI dalla logica
del modello (ordine di intensità), poi si misura. Niente overfitting.
═══════════════════════════════════════════════════════════════════════════

DEFINIZIONI BraTS 2024 (ufficiali):
  Label 1 = NCR (necrotic core)      → ipointenso su T1ce
  Label 2 = ED  (edema / SNFH)       → intensità intermedia
  Label 3 = ET  (enhancing tumor)    → iperintenso su T1ce

  Regioni standard:
    WT = 1+2+3   (whole tumor)
    TC = 1+3     (tumor core = necrosi + enhancing)
    ET = 3       (enhancing)

MAPPING A PRIORI (deciso PRIMA di vedere il ground truth):
  I cluster sono ordinati per intensità della feature primaria crescente
  (cluster 1 = più basso = necrosi; cluster k = più alto = enhancing).
    • ET  → cluster a intensità più ALTA (l'ultimo)
    • TC  → necrosi (cluster 1) + enhancing (cluster k) = primo + ultimo
    • WT  → tutti i cluster (sanity check, =1 per costruzione)
  Questo mapping è biologicamente motivato (enhancing iperintenso su T1ce)
  e NON usa il ground truth. Quindi è onesto.


METRICHE:
  • ARI         — accordo strutturale partizione modello vs 3 classi BraTS
  • Dice TC/ET  — con mapping A PRIORI (non ottimizzato sul GT)
  • Jaccard, HD95 — idem
  • runtime
"""
from __future__ import annotations
import json
import logging
import time
from dataclasses import dataclass, field, asdict
from pathlib import Path

import numpy as np
import nibabel as nib

from io_data.loader import load_patient
from io_data.modality import Modality
from segmentation.registry import get_model

log = logging.getLogger(__name__)


# ── Metriche base (verificate su casi noti) ──────────────────────────────────

def dice_score(pred, gt) -> float:
    pred, gt = pred.astype(bool), gt.astype(bool)
    inter = (pred & gt).sum(); denom = pred.sum() + gt.sum()
    if denom == 0:
        return 1.0 if pred.sum() == 0 else 0.0
    return float(2 * inter / denom)


def jaccard_score(pred, gt) -> float:
    pred, gt = pred.astype(bool), gt.astype(bool)
    inter = (pred & gt).sum(); union = (pred | gt).sum()
    if union == 0:
        return 1.0 if pred.sum() == 0 else 0.0
    return float(inter / union)


def hausdorff_95(pred, gt, spacing=(1,1,1)) -> float:
    from scipy.ndimage import binary_erosion
    from scipy.spatial import KDTree
    pred, gt = pred.astype(bool), gt.astype(bool)
    if pred.sum() == 0 or gt.sum() == 0:
        return float("inf")
    def surf(m): return m & ~binary_erosion(m)
    sp = np.array(spacing)
    p = np.column_stack(np.where(surf(pred))) * sp
    g = np.column_stack(np.where(surf(gt))) * sp
    d1 = KDTree(g).query(p)[0]; d2 = KDTree(p).query(g)[0]
    return float(np.percentile(np.concatenate([d1, d2]), 95))


def apriori_region_masks(label_vol, cluster_ids):
    """
    Mapping A PRIORI cluster→regioni, SENZA guardare il ground truth.

    I cluster_ids sono già ordinati per intensità crescente dalla base class
    (cluster 1 = intensità primaria più bassa, ultimo = più alta).

    Regole fisse basate sulla biologia:
      ET = cluster a intensità più alta (enhancing iperintenso su T1ce)
      TC = necrosi (cluster più basso) + enhancing (cluster più alto)

    Casi limite:
      - 1 solo cluster: ET = TC = quel cluster
      - 2 cluster: ET = alto, TC = entrambi
      - 3+ cluster: ET = alto, TC = primo + ultimo

    Returns: dict con 'TC', 'ET' → (mask, descrizione_cluster)
    """
    if not cluster_ids:
        empty = np.zeros_like(label_vol, dtype=bool)
        return {"TC": (empty, ""), "ET": (empty, "")}

    sorted_ids = sorted(cluster_ids)
    lowest = sorted_ids[0]
    highest = sorted_ids[-1]

    if len(sorted_ids) == 1:
        c = sorted_ids[0]
        mask = (label_vol == c)
        return {"TC": (mask, str(c)), "ET": (mask, str(c))}

    # ET = cluster a più alta intensità
    et_mask = (label_vol == highest)
    et_desc = str(highest)

    # TC = necrosi (più basso) + enhancing (più alto)
    tc_ids = [lowest, highest]
    tc_mask = np.isin(label_vol, tc_ids)
    tc_desc = "+".join(map(str, tc_ids))

    return {"TC": (tc_mask, tc_desc), "ET": (et_mask, et_desc)}


# ── Risultati ─────────────────────────────────────────────────────────────────

@dataclass
class CaseResult:
    patient_id: str
    model_name: str
    dice_wt_sanity: float = 0.0
    dice_tc: float = float("nan")     # NaN finché non valutato (regione presente nel GT)
    dice_et: float = float("nan")
    jaccard_tc: float = float("nan")
    jaccard_et: float = float("nan")
    hd95_tc: float = float("nan")
    hd95_et: float = float("nan")
    ari: float = 0.0
    tc_clusters: str = ""
    et_clusters: str = ""
    n_clusters: int = 0
    gt_has_et: bool = True
    gt_has_tc: bool = True
    runtime_s: float = 0.0
    error: str = ""


@dataclass
class BenchmarkReport:
    model_name: str
    n_cases: int
    n_failed: int
    results: list[CaseResult] = field(default_factory=list)
    mean_dice_tc: float = 0.0; std_dice_tc: float = 0.0; med_dice_tc: float = 0.0
    mean_dice_et: float = 0.0; std_dice_et: float = 0.0; med_dice_et: float = 0.0
    mean_ari: float = 0.0;     std_ari: float = 0.0;     med_ari: float = 0.0
    mean_hd95_tc: float = 0.0; mean_hd95_et: float = 0.0
    mean_runtime: float = 0.0
    n_et_evaluated: int = 0
    n_tc_evaluated: int = 0

    def compute(self):
        valid = [r for r in self.results if not r.error]
        if not valid:
            return
        def stats(vals):
            v = [x for x in vals if x == x and x != float("inf")]
            if not v:
                return 0.0, 0.0, 0.0
            return float(np.mean(v)), float(np.std(v)), float(np.median(v))
        tc_vals = [r.dice_tc for r in valid if r.gt_has_tc]
        et_vals = [r.dice_et for r in valid if r.gt_has_et]
        self.n_tc_evaluated = len(tc_vals)
        self.n_et_evaluated = len(et_vals)
        self.mean_dice_tc, self.std_dice_tc, self.med_dice_tc = stats(tc_vals)
        self.mean_dice_et, self.std_dice_et, self.med_dice_et = stats(et_vals)
        self.mean_ari, self.std_ari, self.med_ari = stats([r.ari for r in valid])
        self.mean_hd95_tc, _, _ = stats([r.hd95_tc for r in valid if r.gt_has_tc])
        self.mean_hd95_et, _, _ = stats([r.hd95_et for r in valid if r.gt_has_et])
        self.mean_runtime, _, _ = stats([r.runtime_s for r in valid])

    def summary(self) -> str:
        self.compute()
        return "\n".join([
            f"━━ BraTS Benchmark [MRI] — {self.model_name} ━━",
            f"Valid cases: {self.n_cases - self.n_failed}/{self.n_cases}",
            "",
            "METHODOLOGY (no metric overfitting):",
            "  • Cluster→region mapping FIXED A PRIORI by intensity,",
            "    WITHOUT looking at the ground truth (ET=highest cluster,",
            "    TC=necrosis+enhancing). No optimization on the GT.",
            "  • Segmentation operates within seg>0 (tumor already localized):",
            "    the internal subdivision is evaluated, not the detection.",
            "  • Regions evaluated only if present in the ground truth.",
            "  • ARI = structural agreement (more robust metric).",
            "",
            "── Results (mean ± std, [median]) ──",
            f"  Dice TC (1+3): {self.mean_dice_tc:.3f} ± {self.std_dice_tc:.3f}  "
            f"[{self.med_dice_tc:.3f}]  (n={self.n_tc_evaluated})",
            f"  Dice ET (3):   {self.mean_dice_et:.3f} ± {self.std_dice_et:.3f}  "
            f"[{self.med_dice_et:.3f}]  (n={self.n_et_evaluated})",
            f"  HD95 TC: {self.mean_hd95_tc:.1f} mm   HD95 ET: {self.mean_hd95_et:.1f} mm",
            "",
            f"  Structural ARI: {self.mean_ari:.3f} ± {self.std_ari:.3f}  "
            f"[{self.med_ari:.3f}]",
            f"  Runtime: {self.mean_runtime:.1f} s/case",
        ])

    def to_excel(self, path):
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Results"
        hf = PatternFill("solid", fgColor="1A3A5C"); hfont = Font(color="FFFFFF", bold=True)
        headers = ["Patient","Model","Dice TC","Dice ET","Jaccard TC","Jaccard ET",
                   "HD95 TC","HD95 ET","ARI","TC=cluster(a priori)","ET=cluster(a priori)",
                   "k","GT has TC","GT has ET","Runtime(s)","WT(sanity)","Error"]
        ws.append(headers)
        for c in ws[1]: c.fill = hf; c.font = hfont
        for r in self.results:
            def fh(v): return round(v,2) if (v==v and v!=float("inf")) else ("—" if v!=v else "inf")
            def fd(v): return round(v,4) if v==v else "—"
            ws.append([r.patient_id, r.model_name, fd(r.dice_tc), fd(r.dice_et),
                       fd(r.jaccard_tc), fd(r.jaccard_et), fh(r.hd95_tc), fh(r.hd95_et),
                       round(r.ari,4), r.tc_clusters, r.et_clusters, r.n_clusters,
                       "yes" if r.gt_has_tc else "no", "yes" if r.gt_has_et else "no",
                       round(r.runtime_s,2), round(r.dice_wt_sanity,3), r.error])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 13
        ws2 = wb.create_sheet("Summary")
        self.compute()
        n_valid = len([r for r in self.results if not r.error])
        ws2.append(["Metric","Mean","Std","Median","N cases"])
        for c in ws2[1]: c.fill = hf; c.font = hfont
        ws2.append(["Dice TC", round(self.mean_dice_tc,4), round(self.std_dice_tc,4),
                    round(self.med_dice_tc,4), self.n_tc_evaluated])
        ws2.append(["Dice ET", round(self.mean_dice_et,4), round(self.std_dice_et,4),
                    round(self.med_dice_et,4), self.n_et_evaluated])
        ws2.append(["ARI", round(self.mean_ari,4), round(self.std_ari,4),
                    round(self.med_ari,4), n_valid])
        ws2.append(["HD95 TC (mm)", round(self.mean_hd95_tc,2), "—","—", self.n_tc_evaluated])
        ws2.append(["HD95 ET (mm)", round(self.mean_hd95_et,2), "—","—", self.n_et_evaluated])
        ws2.append(["Runtime (s)", round(self.mean_runtime,2), "—","—", n_valid])
        ws2.append([])
        for note in [
            "METHODOLOGY (no metric overfitting):",
            "- Cluster->region mapping FIXED A PRIORI (by intensity),",
            "  NOT optimized on the ground truth. No overfitting.",
            "- ET = highest-intensity cluster (hyperintense enhancing on T1ce)",
            "- TC = necrosis (low cluster) + enhancing (high cluster)",
            "- Segmentation within seg>0: subdivision is evaluated, not detection",
            "- Regions evaluated only if present in the GT (NaN excluded from means)",
            "- ARI = structural agreement, more robust metric",
            "- We report mean AND median: the median is more robust to outliers",
        ]:
            ws2.append([note])
        path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        log.info(f"Excel saved: {path}")


class BraTSValidator:
    def __init__(self, brats_dir, output_dir, model_name="GMM",
                 model_params=None, max_cases=None, compute_subregions=True):
        self.brats_dir = Path(brats_dir)
        self.output_dir = Path(output_dir)
        self.model_name = model_name
        self.model_params = model_params or {}
        self.max_cases = max_cases

    def run(self, progress_callback=None) -> BenchmarkReport:
        cases = sorted([d for d in self.brats_dir.iterdir()
                        if d.is_dir() and not d.name.startswith(".")])
        if self.max_cases:
            cases = cases[:self.max_cases]
        report = BenchmarkReport(self.model_name, len(cases), 0)
        log.info(f"HONEST BraTS benchmark (a-priori mapping): {len(cases)} cases, {self.model_name}")
        for i, cd in enumerate(cases):
            if progress_callback:
                progress_callback(i, len(cases), cd.name)
            res = self._process(cd)
            report.results.append(res)
            if res.error:
                report.n_failed += 1
                log.warning(f"[{cd.name}] {res.error}")
            else:
                log.info(f"[{cd.name}] TC={res.dice_tc:.3f} ET={res.dice_et:.3f} ARI={res.ari:.3f}")
        report.compute()
        self.output_dir.mkdir(parents=True, exist_ok=True)
        report.to_excel(self.output_dir / f"brats_MRI_{self.model_name}.xlsx")
        with open(self.output_dir / f"brats_MRI_{self.model_name}.json","w") as f:
            json.dump([asdict(r) for r in report.results], f, indent=2)
        return report

    def _process(self, case_dir) -> CaseResult:
        from sklearn.metrics import adjusted_rand_score
        pid = case_dir.name
        try:
            t0 = time.time()
            data = load_patient(case_dir, force_modality=Modality.MRI)
            features = data.build_features(normalize=True)
            context = data.build_context()
            model = get_model(self.model_name, **self.model_params)
            result = model.fit(features, context)
            runtime = time.time() - t0

            seg = data.volumes["_seg_raw"].astype(np.uint8)
            mask = data.mask
            spacing = tuple(abs(float(data.affine[i,i])) for i in range(3))

            # Definizioni BraTS 2024 corrette (no label 4)
            gt_wt = seg > 0
            gt_tc = (seg == 1) | (seg == 3)
            gt_et = (seg == 3)

            label_vol = result.label_volume
            cluster_ids = [int(c) for c in np.unique(label_vol) if c > 0]

            res = CaseResult(
                patient_id=pid, model_name=self.model_name,
                dice_wt_sanity=dice_score(label_vol > 0, gt_wt),
                n_clusters=len(cluster_ids), runtime_s=runtime,
                gt_has_tc=bool(gt_tc.sum() > 0),
                gt_has_et=bool(gt_et.sum() > 0),
            )

            # ARI strutturale (onesto: non usa GT per il mapping)
            res.ari = float(adjusted_rand_score(
                seg[mask].astype(int), label_vol[mask].astype(int)))

            # MAPPING A PRIORI (deciso senza guardare il GT)
            regions = apriori_region_masks(label_vol, cluster_ids)
            pred_tc, tc_desc = regions["TC"]
            pred_et, et_desc = regions["ET"]

            if res.gt_has_tc:
                res.dice_tc = dice_score(pred_tc, gt_tc)
                res.jaccard_tc = jaccard_score(pred_tc, gt_tc)
                res.hd95_tc = hausdorff_95(pred_tc, gt_tc, spacing)
                res.tc_clusters = tc_desc
            if res.gt_has_et:
                res.dice_et = dice_score(pred_et, gt_et)
                res.jaccard_et = jaccard_score(pred_et, gt_et)
                res.hd95_et = hausdorff_95(pred_et, gt_et, spacing)
                res.et_clusters = et_desc

            return res
        except Exception as e:
            import traceback; traceback.print_exc()
            return CaseResult(pid, self.model_name, error=str(e), hd95_tc=float("inf"))
